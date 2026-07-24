"""대중교통 만족도 조사 XLSX 묶음의 품질 감사와 무손실 정규화.

이 자료는 도시·인구집단별 만족도 평균이며 경로 대안의 선택 라벨이 아니다.
따라서 이 모듈은 점수를 랭커 학습 데이터로 변환하지 않고, 원본 지표명을
연도별로 보존한 감사 결과와 선택적인 정규화 관측값만 만든다.
"""
from __future__ import annotations

import argparse
import hashlib
import io
import json
import math
import re
import zipfile
from collections import Counter
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook

SCHEMA_VERSION = "transit-satisfaction-audit-v1"
SUPPORTED_YEARS = (2023, 2024, 2025)
EXPECTED_REGION_COUNT = 161
MAX_ARCHIVE_BYTES = 50 * 1024 * 1024
MAX_XLSX_BYTES = 15 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 96 * 1024 * 1024
MAX_MEMBER_COMPRESSION_RATIO = 250
MISSING_SCORE_TOKEN = ","
SCORE_RANGE_EPSILON = 1e-9

TITLE_PATTERN = re.compile(
    r"\((?P<year>20\d{2})년\)\s*\[(?P<index>\d+)\.\s*(?P<region>[^\]]+)\]"
)
YEAR_PATTERN = re.compile(r"(?P<year>20\d{2})년")
SEGMENT_GROUPS = {
    "전체": "overall",
    "성별": "gender",
    "연령": "age",
    "주 이용 교통수단": "primary_mode",
    "이용횟수": "usage_frequency",
}


class SurveyDataError(ValueError):
    """조사 파일 구조나 값이 명시된 계약과 다를 때 발생한다."""


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _safe_member_name(name: str) -> bool:
    normalized = name.replace("\\", "/")
    path = PurePosixPath(normalized)
    return (
        bool(normalized)
        and not path.is_absolute()
        and ".." not in path.parts
        and not normalized.startswith("/")
    )


def _validate_xlsx_container(
    payload: bytes,
    source_name: str,
) -> dict[str, bool]:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as workbook_zip:
            entries = workbook_zip.infolist()
            if not entries:
                raise SurveyDataError(f"{source_name}: 빈 XLSX 컨테이너입니다.")
            total_size = 0
            normalized_names: list[str] = []
            for entry in entries:
                if not _safe_member_name(entry.filename):
                    raise SurveyDataError(
                        f"{source_name}: 안전하지 않은 XLSX 내부 경로입니다."
                    )
                total_size += entry.file_size
                if entry.file_size and entry.compress_size == 0:
                    raise SurveyDataError(
                        f"{source_name}: 비정상 압축 항목이 있습니다."
                    )
                if (
                    entry.compress_size
                    and entry.file_size / entry.compress_size
                    > MAX_MEMBER_COMPRESSION_RATIO
                ):
                    raise SurveyDataError(
                        f"{source_name}: 과도한 압축률의 XLSX 항목이 있습니다."
                    )
                normalized_names.append(entry.filename.replace("\\", "/").lower())
            if total_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise SurveyDataError(
                    f"{source_name}: XLSX 압축 해제 크기가 제한을 넘습니다."
                )
            return {
                "unsafe_member_path_present": False,
                "vba_macro_present": any(
                    name.endswith("/vbaproject.bin")
                    for name in normalized_names
                ),
                "external_link_part_present": any(
                    name.startswith("xl/externallinks/")
                    for name in normalized_names
                ),
                "connection_part_present": "xl/connections.xml" in normalized_names,
            }
    except zipfile.BadZipFile as exc:
        raise SurveyDataError(f"{source_name}: 올바른 XLSX가 아닙니다.") from exc


def _workbook_year(name: str) -> int:
    match = YEAR_PATTERN.search(name)
    if match is None:
        raise SurveyDataError(f"{name}: 파일명에서 조사 연도를 찾지 못했습니다.")
    return int(match.group("year"))


def _parse_score(
    value: Any,
    *,
    year: int,
    region: str,
    row_number: int,
    metric: str,
) -> tuple[float | None, str | None]:
    if value is None or value == MISSING_SCORE_TOKEN:
        return None, MISSING_SCORE_TOKEN if value == MISSING_SCORE_TOKEN else "blank"
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise SurveyDataError(
            f"{year} {region} {row_number}행 {metric}: "
            f"허용하지 않은 점수 값 {value!r}"
        )
    numeric = float(value)
    if not math.isfinite(numeric):
        raise SurveyDataError(
            f"{year} {region} {row_number}행 {metric}: 유한하지 않은 점수입니다."
        )
    return numeric, None


def _schema_fingerprint(metric_contract: list[dict[str, str]]) -> str:
    return _sha256(
        json.dumps(
            metric_contract,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    )


def _parse_workbook(
    payload: bytes,
    source_name: str,
    *,
    expected_region_count: int,
    include_observations: bool,
) -> dict[str, Any]:
    container_checks = _validate_xlsx_container(payload, source_name)
    year = _workbook_year(source_name)
    try:
        workbook = load_workbook(
            io.BytesIO(payload),
            # 1MB 안팎의 현재 계약에서는 일반 모드가 read_only 스트리밍보다
            # 훨씬 빠르다. 앞단의 XLSX 압축 해제 크기 제한으로 메모리 상한을
            # 먼저 검증한다.
            read_only=False,
            data_only=True,
        )
    except Exception as exc:
        raise SurveyDataError(f"{source_name}: XLSX를 읽지 못했습니다.") from exc

    try:
        missing_sheets = {"Index", "Table"}.difference(workbook.sheetnames)
        if missing_sheets:
            raise SurveyDataError(
                f"{source_name}: 필수 시트가 없습니다: "
                + ", ".join(sorted(missing_sheets))
            )

        index_sheet = workbook["Index"]
        table_sheet = workbook["Table"]
        # 각 시트를 한 번만 순회해 작은 값 행렬로 고정한 뒤 파싱한다.
        index_rows = list(index_sheet.iter_rows(values_only=True))
        table_rows = list(table_sheet.iter_rows(values_only=True))
        table_column_count = max((len(row) for row in table_rows), default=0)

        def table_value(row_number: int, column_number: int) -> Any:
            row = table_rows[row_number - 1]
            return row[column_number - 1] if column_number <= len(row) else None

        index_regions: list[str] = []
        for row_number, row in enumerate(index_rows[1:], start=2):
            value = row[0] if row else None
            match = TITLE_PATTERN.search(str(value or ""))
            if match is None:
                raise SurveyDataError(
                    f"{source_name}: Index {row_number}행 앵커가 올바르지 않습니다."
                )
            if int(match.group("year")) != year:
                raise SurveyDataError(
                    f"{source_name}: Index의 조사 연도가 파일명과 다릅니다."
                )
            index_regions.append(match.group("region").strip())

        if len(index_regions) != expected_region_count:
            raise SurveyDataError(
                f"{source_name}: 지역 수 {len(index_regions)}개, "
                f"예상 {expected_region_count}개"
            )
        if len(index_regions) != len(set(index_regions)):
            raise SurveyDataError(f"{source_name}: Index 지역명이 중복됩니다.")

        title_rows: list[tuple[int, int, str]] = []
        for row_number, row in enumerate(table_rows, start=1):
            value = row[0] if row else None
            match = TITLE_PATTERN.search(str(value or ""))
            if match is not None:
                title_rows.append(
                    (
                        row_number,
                        int(match.group("index")),
                        match.group("region").strip(),
                    )
                )
        if len(title_rows) != expected_region_count:
            raise SurveyDataError(
                f"{source_name}: Table 지역 블록 수 {len(title_rows)}개, "
                f"예상 {expected_region_count}개"
            )
        table_regions = [region for _, _, region in title_rows]
        if table_regions != index_regions:
            raise SurveyDataError(
                f"{source_name}: Index와 Table의 지역 순서가 다릅니다."
            )

        first_header_row = title_rows[0][0] + 1
        metrics = [
            str(table_value(first_header_row, column) or "").strip()
            for column in range(3, table_column_count + 1)
        ]
        while metrics and not metrics[-1]:
            metrics.pop()
        if not metrics or any(not metric for metric in metrics):
            raise SurveyDataError(f"{source_name}: 지표 헤더가 비어 있습니다.")
        if len(metrics) != len(set(metrics)):
            raise SurveyDataError(f"{source_name}: 지표명이 중복됩니다.")
        metric_contract = [
            {
                "metric_id": f"metric_{index:02d}",
                "source_label": metric,
                "unit_label": str(
                    table_value(first_header_row + 1, index + 2) or ""
                ).strip(),
            }
            for index, metric in enumerate(metrics, start=1)
        ]

        observations: list[dict[str, Any]] = []
        score_values: list[float] = []
        missing_tokens: Counter[str] = Counter()
        duplicate_keys: set[tuple[str, str, str]] = set()
        observed_keys: set[tuple[str, str, str]] = set()
        overall_regions: set[str] = set()
        segment_counts: Counter[str] = Counter()
        observed_segment_values: dict[str, set[str]] = {
            group: set() for group in SEGMENT_GROUPS.values()
        }

        for block_index, (title_row, region_index, region) in enumerate(title_rows):
            expected_index = block_index + 1
            if region_index != expected_index:
                raise SurveyDataError(
                    f"{source_name}: 지역 번호 {region_index}, "
                    f"예상 {expected_index}"
                )
            header_row = title_row + 1
            block_metrics = [
                str(table_value(header_row, column) or "").strip()
                for column in range(3, 3 + len(metrics))
            ]
            if block_metrics != metrics:
                raise SurveyDataError(
                    f"{source_name}: {region} 블록의 지표 스키마가 다릅니다."
                )
            next_title_row = (
                title_rows[block_index + 1][0]
                if block_index + 1 < len(title_rows)
                else len(table_rows) + 1
            )
            current_group: str | None = None
            for row_number in range(title_row + 3, next_title_row):
                category_value = table_value(row_number, 1)
                segment_value = table_value(row_number, 2)
                category = str(category_value or "").strip()
                if category.startswith("BASE"):
                    break
                if category:
                    if category not in SEGMENT_GROUPS:
                        raise SurveyDataError(
                            f"{source_name}: {region} {row_number}행의 "
                            f"알 수 없는 집단 구분 {category!r}"
                        )
                    current_group = SEGMENT_GROUPS[category]
                source_segment = str(segment_value or "").strip()
                if current_group is None and not source_segment:
                    continue
                if current_group is None or not source_segment:
                    raise SurveyDataError(
                        f"{source_name}: {region} {row_number}행의 "
                        "집단 키가 불완전합니다."
                    )
                normalized_segment = (
                    "all" if current_group == "overall" else source_segment
                )
                key = (region, current_group, normalized_segment)
                if key in observed_keys:
                    duplicate_keys.add(key)
                observed_keys.add(key)
                if current_group == "overall":
                    if source_segment != region:
                        raise SurveyDataError(
                            f"{source_name}: {region} 전체행의 지역명이 다릅니다."
                        )
                    overall_regions.add(region)
                segment_counts[current_group] += 1
                observed_segment_values[current_group].add(normalized_segment)

                scores: dict[str, float | None] = {}
                row_missing: dict[str, str] = {}
                for metric_offset, metric in enumerate(metrics, start=3):
                    score, missing_token = _parse_score(
                        table_value(row_number, metric_offset),
                        year=year,
                        region=region,
                        row_number=row_number,
                        metric=metric,
                    )
                    metric_id = f"metric_{metric_offset - 2:02d}"
                    scores[metric_id] = score
                    if score is not None:
                        score_values.append(score)
                    if missing_token is not None:
                        missing_tokens[missing_token] += 1
                        row_missing[metric_id] = missing_token
                if include_observations:
                    observation: dict[str, Any] = {
                        "region_index": region_index,
                        "region": region,
                        "segment_group": current_group,
                        "segment_value": normalized_segment,
                        "source_segment_value": source_segment,
                        "scores": scores,
                    }
                    if row_missing:
                        observation["missing_tokens"] = row_missing
                    observations.append(observation)

        if duplicate_keys:
            sample = sorted(duplicate_keys)[0]
            raise SurveyDataError(
                f"{source_name}: 지역·집단 키가 중복됩니다: {sample}"
            )
        if overall_regions != set(index_regions):
            raise SurveyDataError(
                f"{source_name}: 전체 관측값이 없는 지역이 있습니다."
            )
        if not score_values:
            raise SurveyDataError(f"{source_name}: 유효한 수치 점수가 없습니다.")

        result: dict[str, Any] = {
            "year": year,
            "source_file_name": source_name,
            "source_file_sha256": _sha256(payload),
            "table_rows": len(table_rows),
            "table_columns": table_column_count,
            "region_count": len(index_regions),
            "metric_count": len(metrics),
            "metric_schema_sha256": _schema_fingerprint(metric_contract),
            "metrics": metric_contract,
            "segment_row_counts": dict(sorted(segment_counts.items())),
            "observed_segment_values": {
                group: sorted(values)
                for group, values in sorted(observed_segment_values.items())
            },
            "numeric_score_count": len(score_values),
            "missing_score_count": sum(missing_tokens.values()),
            "missing_tokens": dict(sorted(missing_tokens.items())),
            "observed_min": min(score_values),
            "observed_max": max(score_values),
            "outside_1_to_7_count": sum(
                value < 1 - SCORE_RANGE_EPSILON
                or value > 7 + SCORE_RANGE_EPSILON
                for value in score_values
            ),
            "duplicate_region_segment_keys": 0,
            "xlsx_container_checks": container_checks,
        }
        if include_observations:
            result["observations"] = observations
        return result
    finally:
        workbook.close()


def audit_archive(
    archive_path: Path,
    *,
    include_observations: bool = False,
    expected_region_count: int = EXPECTED_REGION_COUNT,
) -> dict[str, Any]:
    """압축파일 전체를 검증하고 결정적인 감사 JSON 객체를 반환한다."""
    archive_path = archive_path.resolve()
    size = archive_path.stat().st_size
    if size > MAX_ARCHIVE_BYTES:
        raise SurveyDataError("조사 압축파일 크기가 제한을 넘습니다.")
    archive_payload = archive_path.read_bytes()
    try:
        with zipfile.ZipFile(io.BytesIO(archive_payload)) as archive:
            entries = [
                entry
                for entry in archive.infolist()
                if not entry.is_dir()
            ]
            if len(entries) != len(SUPPORTED_YEARS):
                raise SurveyDataError(
                    "압축에는 2023·2024·2025 XLSX 3개만 있어야 합니다."
                )
            workbooks: list[dict[str, Any]] = []
            seen_years: set[int] = set()
            for entry in sorted(entries, key=lambda item: item.filename):
                if (
                    not _safe_member_name(entry.filename)
                    or PurePosixPath(entry.filename).name != entry.filename
                ):
                    raise SurveyDataError("압축 내부 경로가 안전하지 않습니다.")
                if not entry.filename.lower().endswith(".xlsx"):
                    raise SurveyDataError("압축에는 XLSX 파일만 허용됩니다.")
                if entry.file_size > MAX_XLSX_BYTES:
                    raise SurveyDataError(
                        f"{entry.filename}: 파일 크기가 제한을 넘습니다."
                    )
                payload = archive.read(entry)
                year = _workbook_year(entry.filename)
                if year in seen_years:
                    raise SurveyDataError(f"{year}년 파일이 중복됩니다.")
                seen_years.add(year)
                workbooks.append(
                    _parse_workbook(
                        payload,
                        entry.filename,
                        expected_region_count=expected_region_count,
                        include_observations=include_observations,
                    )
                )
    except zipfile.BadZipFile as exc:
        raise SurveyDataError("올바른 ZIP 압축파일이 아닙니다.") from exc

    if seen_years != set(SUPPORTED_YEARS):
        raise SurveyDataError("2023·2024·2025 조사 파일이 모두 필요합니다.")
    workbooks.sort(key=lambda item: item["year"])
    schema_counts = {item["year"]: item["metric_count"] for item in workbooks}

    def all_years_have(group: str, *values: str) -> bool:
        return all(
            set(values).issubset(
                set(workbook["observed_segment_values"].get(group, []))
            )
            for workbook in workbooks
        )

    profile_context_coverage = {
        "general": {
            "status": (
                "city_context_only"
                if all_years_have("overall", "all")
                else "absent"
            ),
            "source_segments": ["overall:all"],
            "direct_route_label": False,
        },
        "elderly": {
            "status": (
                "city_context_only"
                if all_years_have("age", "60대 이상")
                else "absent"
            ),
            "source_segments": ["age:60대 이상"],
            "direct_route_label": False,
        },
        "youth": {
            "status": (
                "mapping_undefined"
                if all_years_have("age", "10대", "20대")
                else "absent"
            ),
            "source_segments": ["age:10대", "age:20대"],
            "direct_route_label": False,
        },
        "child": {
            "status": "no_direct_segment",
            "source_segments": [],
            "direct_route_label": False,
        },
        "disabled": {
            "status": "no_direct_segment",
            "source_segments": [],
            "direct_route_label": False,
            "note": "교통약자 시설은 서비스 문항이며 개인 집단이 아닙니다.",
        },
        "pregnant": {
            "status": "no_direct_segment",
            "source_segments": [],
            "direct_route_label": False,
        },
        "carry_luggage": {
            "status": "no_direct_segment",
            "source_segments": [],
            "direct_route_label": False,
        },
    }
    warnings = [
        "원본 파일 안에서 조사기관·표본수·가중치·척도·라이선스 정보를 확인할 수 없습니다.",
        "연도별 지표 스키마가 달라 지표명 확인 없는 단순 시계열 결합을 금지합니다.",
        "이 자료는 도시·집단별 평균이며 OD·경로 대안 선호 라벨이 아닙니다.",
    ]
    if any(item["outside_1_to_7_count"] for item in workbooks):
        warnings.append(
            "관측값 중 1~7 범위를 벗어난 값이 있어 원 조사 산식 확인 전 척도를 단정할 수 없습니다."
        )

    result: dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "source_archive_name": archive_path.name,
        "source_archive_sha256": _sha256(archive_payload),
        "source_archive_bytes": size,
        "dataset_scope": {
            "years": list(SUPPORTED_YEARS),
            "region_count_per_year": expected_region_count,
            "score_level": "region_by_aggregate_segment_mean",
        },
        "audit_scope": {
            "parsed_sheets": ["Index", "Table"],
            "xlsx_part_names_checked": True,
            "formula_semantics_checked": False,
            "respondent_level_pii_scan_applicable": False,
        },
        "metric_count_by_year": schema_counts,
        "workbooks": workbooks,
        "model_use_policy": {
            "route_ranking_label": "prohibited",
            "global_or_profile_weight": "requires_methodology_and_policy_review",
            "current_allowed_use": [
                "data_quality_audit",
                "coverage_gap_analysis",
                "city_level_context_after_license_confirmation",
                "feedback_question_taxonomy_non_model",
                "future_external_validity_comparison",
            ],
        },
        "profile_context_coverage": profile_context_coverage,
        "provenance_status": {
            "source_url": None,
            "license": None,
            "methodology": None,
            "redistribution_allowed": None,
        },
        "warnings": warnings,
    }
    return result


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, allow_nan=False) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="3개년 대중교통 만족도 조사 ZIP을 감사·정규화합니다."
    )
    parser.add_argument("--archive", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--include-observations",
        action="store_true",
        help="라이선스 확인 후에만 실제 점수 관측값을 출력합니다.",
    )
    parser.add_argument(
        "--acknowledge-license-reviewed",
        action="store_true",
        help="원자료 이용·재배포 조건을 검토했다는 명시적 확인입니다.",
    )
    args = parser.parse_args()
    if args.include_observations and not args.acknowledge_license_reviewed:
        parser.error(
            "--include-observations에는 "
            "--acknowledge-license-reviewed 확인이 필요합니다."
        )
    result = audit_archive(
        args.archive,
        include_observations=args.include_observations,
    )
    _write_json(args.output, result)


if __name__ == "__main__":
    main()

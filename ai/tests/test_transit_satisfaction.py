from __future__ import annotations

import io
import json
import os
import time
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from data_tools.transit_satisfaction import (
    SurveyDataError,
    audit_archive,
)

REAL_ARCHIVE = os.getenv("TRANSIT_SURVEY_ARCHIVE")


def _workbook_bytes(year: int, regions: list[str]) -> bytes:
    workbook = Workbook()
    index = workbook.active
    index.title = "Index"
    table = workbook.create_sheet("Table")
    index.append(["Table of Contents", None])
    table_row = 1
    for region_index, region in enumerate(regions, start=1):
        title = (
            f"<만족도 01> 만족도 테이블 전체({year}년) "
            f"[{region_index}. {region}]"
        )
        index.append([title, "Index"])
        table.cell(table_row, 1, title)
        table.cell(table_row + 1, 3, "전반적 만족도")
        table.cell(table_row + 1, 4, "교통약자 시설")
        table.cell(table_row + 2, 3, "평균(점)")
        table.cell(table_row + 2, 4, "평균(점)")
        table.cell(table_row + 3, 1, "전체")
        table.cell(table_row + 3, 2, region)
        table.cell(table_row + 3, 3, 4.5 + region_index / 10)
        table.cell(table_row + 3, 4, ",")
        table.cell(table_row + 4, 1, "연령")
        table.cell(table_row + 4, 2, "60대 이상")
        table.cell(table_row + 4, 3, 4.2)
        table.cell(table_row + 4, 4, 4.1)
        table.cell(table_row + 5, 1, "BASE : 전체")
        table_row += 7
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _archive(tmp_path: Path, *, unsafe_name: str | None = None) -> Path:
    archive_path = tmp_path / "survey.zip"
    with zipfile.ZipFile(archive_path, "w") as archive:
        for year in (2023, 2024, 2025):
            name = unsafe_name if year == 2023 and unsafe_name else f"만족도_{year}년.xlsx"
            archive.writestr(name, _workbook_bytes(year, ["서울", "부산"]))
    return archive_path


def test_audit_preserves_missing_values_and_prohibits_route_label_use(
    tmp_path: Path,
):
    result = audit_archive(
        _archive(tmp_path),
        include_observations=True,
        expected_region_count=2,
    )

    assert result["metric_count_by_year"] == {2023: 2, 2024: 2, 2025: 2}
    assert result["model_use_policy"]["route_ranking_label"] == "prohibited"
    assert result["provenance_status"]["license"] is None
    first_year = result["workbooks"][0]
    assert first_year["missing_score_count"] == 2
    assert first_year["missing_tokens"] == {",": 2}
    overall = first_year["observations"][0]
    assert overall["scores"]["metric_02"] is None
    assert overall["missing_tokens"]["metric_02"] == ","


def test_audit_rejects_unsafe_archive_member_path(tmp_path: Path):
    with pytest.raises(SurveyDataError, match="안전하지"):
        audit_archive(
            _archive(tmp_path, unsafe_name="../만족도_2023년.xlsx"),
            expected_region_count=2,
        )


def test_audit_rejects_unknown_score_text(tmp_path: Path):
    archive_path = tmp_path / "survey.zip"
    payloads = {
        year: _workbook_bytes(year, ["서울", "부산"])
        for year in (2023, 2024, 2025)
    }
    workbook = load_workbook(io.BytesIO(payloads[2024]))
    workbook["Table"].cell(4, 3, "알 수 없음")
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    payloads[2024] = buffer.getvalue()
    with zipfile.ZipFile(archive_path, "w") as archive:
        for year, payload in payloads.items():
            archive.writestr(f"만족도_{year}년.xlsx", payload)

    with pytest.raises(SurveyDataError, match="허용하지 않은 점수"):
        audit_archive(archive_path, expected_region_count=2)


@pytest.mark.skipif(
    not REAL_ARCHIVE,
    reason="실제 조사 ZIP은 라이선스 확인 전 저장소에 포함하지 않습니다.",
)
def test_real_archive_golden_contract_and_runtime():
    started = time.perf_counter()
    result = audit_archive(Path(str(REAL_ARCHIVE)))
    elapsed = time.perf_counter() - started

    assert elapsed < 30
    assert result["source_archive_sha256"] == (
        "24f6634560867063001da25b1cd5a639"
        "e4542e2d0b2b7a8f6158321c08e9de12"
    )
    assert result["metric_count_by_year"] == {
        2023: 29,
        2024: 21,
        2025: 23,
    }
    assert [
        (
            workbook["numeric_score_count"],
            workbook["missing_score_count"],
            workbook["outside_1_to_7_count"],
        )
        for workbook in result["workbooks"]
    ] == [
        (70501, 607, 55),
        (51018, 474, 0),
        (52098, 3677, 0),
    ]


def test_committed_audit_summary_matches_golden_contract():
    audit_path = (
        Path(__file__).resolve().parents[2]
        / "data"
        / "audits"
        / "public_transport_satisfaction_2023_2025.audit.json"
    )
    summary = json.loads(audit_path.read_text(encoding="utf-8"))

    assert summary["sourceArchive"]["retainedInRepository"] is False
    assert summary["sourceArchive"]["sha256"] == (
        "24f6634560867063001da25b1cd5a639"
        "e4542e2d0b2b7a8f6158321c08e9de12"
    )
    assert [
        (
            workbook["year"],
            workbook["metricCount"],
            workbook["numericScoreCount"],
            workbook["missingCommaCount"],
        )
        for workbook in summary["workbooks"]
    ] == [
        (2023, 29, 70501, 607),
        (2024, 21, 51018, 474),
        (2025, 23, 52098, 3677),
    ]
    assert summary["modelUsePolicy"]["routeRankingLabel"] == "prohibited"
    assert summary["provenanceStatus"]["license"] is None
